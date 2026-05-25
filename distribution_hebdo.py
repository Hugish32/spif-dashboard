import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os
from datetime import datetime

CSV_FILE = "Pourcentage_produit_par_region___claude.csv"
WHITE  = "FFFFFF"
GREY1  = "F5F5F5"
GREY2  = "E0E0E0"
GREY3  = "BBBBBB"
GREY4  = "595959"
BLACK  = "000000"
SEP    = "AAAAAA"

CATEGORIES = {
    "Produits Quebec": ["Domaine Les Rassembleurs l'Orange 2020","Aupale Vodka","BleuRoyal Gin","BleuRoyal Pêche","Jardin Verde Gin","Jardin Verde Glam","Jardin Verde Gin Tonic Zero (4 x 355 Ml )","Bravo Charlie","Echo Foxtrot","Mayday Épinette Cannelle","Mission Kosmos","Mugo Gin","Mugo Édition Citrus Mitis","Moonlight","Ciel ! Camerise Mandarine","Aléa Espresso","BluePearl Tropical","Boîte Noire Sambuca Noire","Distillerie Mitis Acerum Vieilli en baril","Jardin Verde Mojito Zero","Echo Foxtrot Smash Basilic Framboise","Silly Billy Cornichons","Première Classe","Moonlight Rhum & Soda Gingembre","Moonlight Crème","Melon Drive Rhum Melon D'eau","Mate Libre Mojito","Kanon Rhum Épicé","Distillerie Mitis Entaille","Bobba Thé aux Perles Limonade Fruit du Dragon","BluePearl Collection, Sureau","Aupale Seltzer Pamplemousse","Rose Crème, Crème de Tequila à la Fraise","UPPé Litchi Martini","UPPé Spicy Margarita"],
    "ALTER": ["Sueno de Alden Mezcal Joven","Sueno de Alden Mezcal Reposado","Knappogue Castle 12 years old Single Malt","Ti Arrangés de Ced Ananas Victoria Rhum","Stock Lionello Rouge","Stock Lionello Blanc","Stock 84","Parasol Lime & cactus","Midi Tequila Blanco","Good Day Pêche","Good Day Melon Soju","Good Day Litchi Soju","Good Day Ananas Soju","Angostura Cinq Ans","Avion Reserva 44 Extra Anejo","Good Day Pêche"],
    "Chateau de Cartes": ["Château de Cartes Vin de Soif","Château de Cartes Vin Orange 2023","Château de Cartes Vin Gris 2022","Château de Cartes Saint-Pépin 2023","Château de Cartes PetNat Rouge 2023","Château de Cartes PetNat Rosé","Château de Cartes PetNat 2022","Château de Cartes Perle Noire 2021","Château de Cartes Marquette Réserve 2021","Château de Cartes Les Affleurements 2021","Château de Cartes Boisé","Château de Cartes Blanc de Noir 2022","Château de Cartes Atout Rouge 2021","Château de Cartes Atout Rosé 2023","Château de Cartes Atout Blanc 2023"],
    "Labatt": ["Cidrerie Lacroix Anne Série Héritage","Corona Sunbrew","Cidrerie Lacroix Sangria d'Ici","Palm Bay Ananas Mandarine","Palm Bay Frozen Rainbow Twist","SVNS Vodka 7UP Berry Mixte","SVNS Hard 7UP Caisse Mixte","Palm Bay Mixer","Cidrerie Lacroix Série Nature Osmose","Cidre Lacroix Margarita d'Ici"],
}

def load_data():
    df = pd.read_csv(CSV_FILE)
    df["DATE DU JOUR"] = pd.to_datetime(df["DATE DU JOUR"]).dt.strftime("%Y-%m-%d")
    df = df.rename(columns={"Représentant":"Representant"})
    df["pct"] = pd.to_numeric(df["Succ avec stock"],errors="coerce") / pd.to_numeric(df["Total Succ"],errors="coerce")
    df = df[~df["Representant"].str.contains("Hugo Carignan",na=False)]
    return df

def find_closest(dates, target):
    target_dt = datetime.strptime(target, "%Y-%m-%d")
    return min(dates, key=lambda d: abs(datetime.strptime(d, "%Y-%m-%d") - target_dt))

def auto_dates(df):
    dates = sorted(df["DATE DU JOUR"].unique())
    dc = dates[-1]; dp = dates[-2] if len(dates) > 1 else dc
    dates_2025 = [d for d in dates if d.startswith("2025")]
    target = datetime.strptime(dc, "%Y-%m-%d").replace(year=2025).strftime("%Y-%m-%d")
    d2 = find_closest(dates_2025, target) if dates_2025 else None
    return dc, dp, d2

def thin():
    s = Side(border_style="thin", color="DDDDDD")
    return Border(left=s,right=s,top=s,bottom=s)

def get_pct(data, prod):
    sub = data[data["Nom Produit"]==prod]
    return sub["pct"].mean() if not sub.empty else float("nan")

def get_succ(data, prod):
    sub = data[(data["Nom Produit"]==prod) & (~data["Representant"].str.contains("Hugo", na=False))]
    if sub.empty: return float("nan")
    pct = pd.to_numeric(sub["pct"], errors="coerce").mean()
    if pd.isna(pct): return float("nan")
    return round(pct * 413)

def get_avail(cat_prods, ref):
    seen = set(); result = []
    for p in cat_prods:
        if p in ref["Nom Produit"].values and p not in seen:
            seen.add(p); result.append(p)
    return result

def cell(ws, row, col, val, fmt=None, bg=WHITE, bold=False, fg=BLACK, size=9, align="center", wrap=False):
    c = ws.cell(row, col, val)
    c.font = Font(name="Arial", size=size, color=fg, bold=bold)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin()
    if fmt: c.number_format = fmt
    return c

def delta(ws, row, col, val, bg=WHITE, color=True, is_succ=False):
    if pd.isna(val) or val is None:
        cell(ws, row, col, "-", bg=bg); return
    fmt = "+0;-0;-" if is_succ else "+0.0%;-0.0%;-"
    c = cell(ws, row, col, val, fmt=fmt, bg=bg)
    if color:
        thresh = 0.5 if is_succ else 0.005
        if val > thresh:
            c.fill = PatternFill("solid", start_color="C6EFCE")
            c.font = Font(name="Arial", size=9, color="375623", bold=True)
        elif val < -thresh:
            c.fill = PatternFill("solid", start_color="FFC7CE")
            c.font = Font(name="Arial", size=9, color="9C0006", bold=True)

def sep_line(ws, row, max_col):
    for col in range(1, max_col+1):
        c = ws.cell(row, col, "")
        c.fill = PatternFill("solid", start_color=GREY3)
    ws.row_dimensions[row].height = 4

def blank_line(ws, row, max_col):
    for col in range(1, max_col+1):
        c = ws.cell(row, col, "")
        c.fill = PatternFill("solid", start_color=WHITE)
        c.border = thin()
    ws.row_dimensions[row].height = 6

def build_col_map(cats_avail):
    col_map = {}; sum_cols = {}; sep_cols = []; col = 2
    for cat, prods in cats_avail.items():
        for p in prods:
            col_map[p] = col; col += 1
        sum_cols[cat] = col; col += 1
        sep_cols.append(col); col += 1
    return col_map, sum_cols, sep_cols

def write_headers(ws, r1, r2, cats_avail, col_map, sum_cols, sep_cols):
    cell(ws, r1, 1, "", bg=GREY4)
    cell(ws, r2, 1, "", bg=GREY4)
    for ci, (cat, prods) in enumerate(cats_avail.items()):
        fc = col_map[prods[0]]
        lc = col_map[prods[-1]]
        sc = sum_cols[cat]
        sp = sep_cols[ci]
        ws.merge_cells(start_row=r1, start_column=fc, end_row=r1, end_column=lc)
        c = ws.cell(r1, fc, cat)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=GREY4)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        for p in prods:
            col = col_map[p]
            c2 = ws.cell(r2, col, p)
            c2.font = Font(bold=True, color=WHITE, name="Arial", size=8)
            c2.fill = PatternFill("solid", start_color=GREY4)
            c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c2.border = thin()
            ws.column_dimensions[get_column_letter(col)].width = 7
        cell(ws, r1, sc, "", bg=GREY4)
        c3 = ws.cell(r2, sc, "Moy.")
        c3.font = Font(bold=True, color=BLACK, name="Arial", size=9)
        c3.fill = PatternFill("solid", start_color=GREY2)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border = thin()
        ws.column_dimensions[get_column_letter(sc)].width = 8
        for r in [r1, r2]:
            ws.cell(r, sp, "").fill = PatternFill("solid", start_color=SEP)
        ws.column_dimensions[get_column_letter(sp)].width = 2
    ws.row_dimensions[r1].height = 18
    ws.row_dimensions[r2].height = 42

def section_header(ws, row, max_col, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row, 1, title)
    c.font = Font(bold=True, color=BLACK, name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=GREY2)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin()
    ws.row_dimensions[row].height = 18

def build_sheet(wb, sheet_name, df, dfc, dfp, df2, dc, dp, d2, rep=None):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.column_dimensions["A"].width = 24
    reps = sorted(dfc["Representant"].dropna().unique().tolist())
    cats_avail = {}
    for cat, prods in CATEGORIES.items():
        avail = get_avail(prods, dfc)
        if avail: cats_avail[cat] = avail
    col_map, sum_cols, sep_cols = build_col_map(cats_avail)
    max_col = sep_cols[-1]

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(1, 1, f"Distribution Hebdomadaire  |  {sheet_name}  |  Semaine du {dc}")
    c.font = Font(bold=True, color=WHITE, name="Arial", size=12)
    c.fill = PatternFill("solid", start_color=GREY4)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    ws.row_dimensions[1].height = 26

    write_headers(ws, 2, 3, cats_avail, col_map, sum_cols, sep_cols)
    row = 4

    # MACRO
    section_header(ws, row, max_col, "MACRO - Nombre de succursales par produit")
    row += 1

    macro_rows = [
        (f"Succursales {dc}", "sc",    WHITE, False, True),
        (f"Succursales {dp}", "sp",    GREY1, False, False),
        ("Variation sem/sem", "ds",    WHITE, True,  False),
        ("",                  "blank", WHITE, False, False),
        (f"Succ {d2 or '2025'}", "s2", GREY1, False, False),
        (f"Variation vs {(d2 or '')[:4]}", "dy", WHITE, True, False),
    ]

    for label, key, bg, color, bold in macro_rows:
        if key == "blank":
            blank_line(ws, row, max_col); row += 1; continue
        cell(ws, row, 1, label, bg=bg, bold=bold, align="left", size=10)
        ws.row_dimensions[row].height = 17

        for ci, (cat, prods) in enumerate(cats_avail.items()):
            vals_sc = []; vals_sp = []
            vals = []
            for p in prods:
                col = col_map[p]
                sv  = get_succ(dfc, p)
                spv = get_succ(dfp, p)
                s2v = get_succ(df2, p) if not df2.empty else float("nan")
                if key == "sc":
                    cell(ws, row, col, sv if not pd.isna(sv) else "-", fmt="0", bg=bg, bold=bold)
                    if not pd.isna(sv): vals.append(sv)
                elif key == "sp":
                    cell(ws, row, col, spv if not pd.isna(spv) else "-", fmt="0", bg=bg)
                    if not pd.isna(spv): vals.append(spv)
                elif key == "ds":
                    v = sv-spv if not pd.isna(sv) and not pd.isna(spv) else float("nan")
                    delta(ws, row, col, v, bg=bg, color=color, is_succ=True)
                    if not pd.isna(sv):  vals_sc.append(sv)
                    if not pd.isna(spv): vals_sp.append(spv)
                elif key == "s2":
                    cell(ws, row, col, s2v if not pd.isna(s2v) else "-", fmt="0", bg=bg)
                    if not pd.isna(s2v): vals.append(s2v)
                elif key == "dy":
                    v = sv-s2v if not pd.isna(sv) and not pd.isna(s2v) else float("nan")
                    delta(ws, row, col, v, bg=bg, color=color, is_succ=True)
                    if not pd.isna(sv):  vals_sc.append(sv)
                    if not pd.isna(s2v): vals_sp.append(s2v)

            moy_col = sum_cols[cat]
            sp_col  = sep_cols[ci]

            if key in ("ds", "dy"):
                moy_a = round(sum(vals_sc)/len(vals_sc)) if vals_sc else None
                moy_b = round(sum(vals_sp)/len(vals_sp)) if vals_sp else None
                if moy_a is not None and moy_b is not None:
                    delta(ws, row, moy_col, moy_a-moy_b, bg=GREY2, color=color, is_succ=True)
                    ws.cell(row, moy_col).font = Font(name="Arial", size=9, bold=True,
                        color=ws.cell(row,moy_col).font.color.rgb
                        if hasattr(ws.cell(row,moy_col).font.color,"rgb")
                        and ws.cell(row,moy_col).font.color.type=="rgb" else BLACK)
                else:
                    cell(ws, row, moy_col, "-", bg=GREY2, bold=True)
            elif key in ("sc","sp","s2") and vals:
                moy = round(sum(vals)/len(vals))
                cell(ws, row, moy_col, moy, fmt="0", bg=GREY2, bold=True)
            else:
                cell(ws, row, moy_col, "-", bg=GREY2, bold=True)

            ws.cell(row, sp_col, "").fill = PatternFill("solid", start_color=SEP)
        row += 1

    sep_line(ws, row, max_col); row += 2

    # PAR REP
    section_header(ws, row, max_col, "PAR REPRESENTANT - Distribution en pourcentage")
    row += 1

    target_reps = [rep] if rep else reps
    for ri, cur_rep in enumerate(target_reps):
        dfc_cr = dfc[dfc["Representant"]==cur_rep]
        dfp_cr = dfp[dfp["Representant"]==cur_rep]
        df2_cr = df2[df2["Representant"]==cur_rep] if not df2.empty else pd.DataFrame()

        # Nom rep
        cell(ws, row, 1, cur_rep, bg=GREY2, bold=True, align="left", size=10)
        for ci, (cat, prods) in enumerate(cats_avail.items()):
            for p in prods:
                cell(ws, row, col_map[p], "", bg=GREY2)
            cell(ws, row, sum_cols[cat], "", bg=GREY2, bold=True)
            ws.cell(row, sep_cols[ci], "").fill = PatternFill("solid", start_color=SEP)
        ws.row_dimensions[row].height = 16; row += 1

        if rep:
            rep_rows = [
                (cur_rep.split()[0]+" - %",        "pct", WHITE, False, True),
                ("Variation sem/sem",               "ds",  GREY1, True,  False),
                (f"Variation vs {(d2 or '')[:4]}", "dy",  WHITE, False, False),
                ("Variation vs equipe",             "deq", GREY1, False, False),
            ]
        else:
            rep_rows = [
                (cur_rep.split()[0]+" - %", "pct", WHITE, False, True),
                ("Variation sem/sem",        "ds",  GREY1, True,  False),
            ]

        for label, key, bg, color, bold in rep_rows:
            cell(ws, row, 1, label, bg=bg, bold=bold, align="left", size=10)
            ws.row_dimensions[row].height = 17

            for ci, (cat, prods) in enumerate(cats_avail.items()):
                vals = []
                for p in prods:
                    col = col_map[p]
                    vr  = get_pct(dfc_cr, p)
                    vpr = get_pct(dfp_cr, p)
                    v2r = get_pct(df2_cr, p) if not df2_cr.empty else float("nan")
                    veq = get_pct(dfc, p)
                    if key == "pct":
                        cell(ws, row, col, vr if not pd.isna(vr) else "-", fmt="0.0%", bg=bg, bold=bold)
                        if not pd.isna(vr): vals.append(vr)
                    elif key == "ds":
                        v = vr-vpr if not pd.isna(vr) and not pd.isna(vpr) else float("nan")
                        delta(ws, row, col, v, bg=bg, color=color)
                        if not pd.isna(v): vals.append(v)
                    elif key == "dy":
                        v = vr-v2r if not pd.isna(vr) and not pd.isna(v2r) else float("nan")
                        delta(ws, row, col, v, bg=bg, color=False)
                        if not pd.isna(v): vals.append(v)
                    elif key == "deq":
                        v = vr-veq if not pd.isna(vr) and not pd.isna(veq) else float("nan")
                        delta(ws, row, col, v, bg=bg, color=False)
                        if not pd.isna(v): vals.append(v)

                moy_col = sum_cols[cat]
                avg = sum(vals)/len(vals) if vals else float("nan")
                if key == "pct":
                    c2 = ws.cell(row, moy_col, avg if not pd.isna(avg) else "-")
                    c2.number_format = "0.0%"
                    c2.font = Font(name="Arial", size=9, bold=True, color=BLACK)
                    c2.fill = PatternFill("solid", start_color=GREY2)
                    c2.alignment = Alignment(horizontal="center", vertical="center")
                    c2.border = thin()
                elif key == "ds":
                    delta(ws, row, moy_col, avg if not pd.isna(avg) else float("nan"), bg=GREY2, color=color)
                    c2 = ws.cell(row, moy_col)
                    if c2.font.color.type == "rgb":
                        c2.font = Font(name="Arial", size=9, bold=True, color=c2.font.color.rgb)
                    else:
                        c2.font = Font(name="Arial", size=9, bold=True, color=BLACK)
                else:
                    delta(ws, row, moy_col, avg if not pd.isna(avg) else float("nan"), bg=GREY2, color=False)
                    ws.cell(row, moy_col).font = Font(name="Arial", size=9, bold=True, color=BLACK)

                ws.cell(row, sep_cols[ci], "").fill = PatternFill("solid", start_color=SEP)
            row += 1

        if ri < len(target_reps)-1:
            sep_line(ws, row, max_col); row += 2

    ws.freeze_panes = "B4"

def main():
    print("="*60)
    print("  Distribution Hebdomadaire")
    print("="*60)
    if not os.path.exists(CSV_FILE):
        print(f"ERREUR: {CSV_FILE} introuvable!"); sys.exit(1)
    df = load_data()
    dc, dp, d2 = auto_dates(df)
    print(f"  Semaine courante   : {dc}")
    print(f"  Semaine precedente : {dp}")
    print(f"  Meme semaine 2025  : {d2}")
    dfc = df[df["DATE DU JOUR"]==dc]
    dfp = df[df["DATE DU JOUR"]==dp]
    df2 = df[df["DATE DU JOUR"]==d2] if d2 else pd.DataFrame()
    reps = sorted(dfc["Representant"].dropna().unique().tolist())
    wb = Workbook(); wb.remove(wb.active)
    build_sheet(wb,"Recap Global",df,dfc,dfp,df2,dc,dp,d2)
    print("  Feuille : Recap Global")
    for rep in reps:
        build_sheet(wb,rep.split()[0],df,dfc,dfp,df2,dc,dp,d2,rep=rep)
        print(f"  Feuille : {rep.split()[0]}")
    fn = f"Distribution_{dc}.xlsx"
    wb.save(fn)
    print(f"  Fichier : {fn}")
    print("="*60)

if __name__ == "__main__":
    main()
