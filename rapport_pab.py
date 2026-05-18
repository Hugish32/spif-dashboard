"""
RAPPORT HEBDOMADAIRE PAB — Labatt
Lit le CSV exporté depuis Zoho Analytics
Version 3 — Cumul correct, page Marques, page Segments
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

FICHIER_CSV    = "VENTES_PAB_ETE_2025_2026.csv"
MARQUES_LABATT = ['PALM BAY', 'SVNS', 'BEACH DAY EVERY DAY']
SEUIL_INNO     = 15500000

BLEU_FONCE  = "1F3864"
BLEU_CLAIR  = "D6E4F0"
VERT_FONCE  = "2E7D32"
VERT_CLAIR  = "E8F5E9"
ROUGE_FONCE = "B71C1C"
ROUGE_CLAIR = "FFEBEE"
ORANGE_FONCE= "BF360C"
ORANGE_CLAIR= "FBE9E7"
VIOLET_FONCE= "4A148C"
VIOLET_CLAIR= "F3E5F5"
GRIS_CLAIR  = "F5F5F5"
GRIS_HEADER = "37474F"

FMT_DOLLAR = '#,##0.00\\ "$"'
FMT_NOMBRE = '#,##0'

# ══════════════════════════════════════════════════════════════════════════════
# STYLE
# ══════════════════════════════════════════════════════════════════════════════

def style_header(cell, bg=None):
    cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg or BLEU_FONCE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_total(cell):
    cell.font = Font(bold=True, size=10, name="Arial")
    cell.fill = PatternFill("solid", start_color=GRIS_CLAIR)
    cell.alignment = Alignment(vertical="center")

def style_data(cell, bold=False, bg=None, align="left"):
    cell.font      = Font(bold=bold, size=10, name="Arial")
    cell.alignment = Alignment(vertical="center", horizontal=align)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def add_border(cell):
    s = Side(style="thin", color="E0E0E0")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, max_w))

def pct_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "N/A"
    return f"{val:+.1f}%"

def titre_page(ws, texte, couleur, nb_cols):
    ws.merge_cells(f"A1:{get_column_letter(nb_cols)}1")
    ws["A1"] = texte
    ws["A1"].font = Font(bold=True, size=14, color=couleur, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

# ══════════════════════════════════════════════════════════════════════════════
# CHARGEMENT ET PRÉPARATION
# ══════════════════════════════════════════════════════════════════════════════

def charger_donnees():
    df = pd.read_csv(FICHIER_CSV)
    df.columns = df.columns.str.strip()
    rename = {c: c[2:] for c in df.columns if c.startswith('v.') or c.startswith('p.')}
    df = df.rename(columns=rename)
    print(f"   Colonnes : {list(df.columns)}")
    print(f"   Lignes   : {len(df):,}")
    return df

def preparer(df):
    col_code  = [c for c in df.columns if 'SAQ CODE' in c][0]
    col_annee = [c for c in df.columns if 'ANNÉE' in c or 'YEAR' in c.upper()][0]
    col_aps   = [c for c in df.columns if 'Annee-Periode' in c][0]
    col_qty   = [c for c in df.columns if 'Bouteilles' in c or 'BOTTLES' in c][-1]
    col_amt   = [c for c in df.columns if 'Montant' in c or 'MONTANT' in c or 'AMOUNT' in c][-1]
    col_desc  = [c for c in df.columns if 'DESCRIPTION' in c.upper()]
    col_marq  = [c for c in df.columns if 'Marque' in c]
    col_seg   = [c for c in df.columns if 'Segment' in c]

    df["Code_SAQ"]   = df[col_code].astype(str).str.strip()
    df["Annee"]      = pd.to_numeric(df[col_annee], errors="coerce")
    df["APS"]        = df[col_aps].astype(str).str.strip()
    df["Quantite"]   = pd.to_numeric(df[col_qty], errors="coerce").fillna(0)
    df["Montant"]    = pd.to_numeric(df[col_amt], errors="coerce").fillna(0)
    df["Description"]= df[col_desc[0]].astype(str).str.strip() if col_desc else ""
    df["Marque_PaB"] = df[col_marq[0]].astype(str).str.strip() if col_marq else ""
    df["Segment"]    = df[col_seg[0]].astype(str).str.strip() if col_seg else ""
    df["Marque_PaB"] = df["Marque_PaB"].replace("nan", "")
    df["Segment"]    = df["Segment"].replace("nan", "")

    # Dernière semaine 2026
    semaines_2026 = sorted(df[df["Annee"] == 2026]["APS"].dropna().unique())
    derniere      = semaines_2026[-1]
    parts         = derniere.split("-")
    derniere_ly   = f"{int(parts[0])-1}-{parts[1]}-{parts[2]}"
    print(f"   Dernière semaine : {derniere}  |  An passé : {derniere_ly}")

    # ── Semaine courante ──────────────────────────────────────────────
    df_s = df[df["APS"] == derniere].groupby("Code_SAQ").agg(
        Quantite    =("Quantite",    "sum"),
        Montant     =("Montant",     "sum"),
        Description =("Description", "first"),
        Marque_PaB  =("Marque_PaB",  "first"),
        Segment     =("Segment",     "first"),
    ).reset_index()

    # ── Même semaine an passé ─────────────────────────────────────────
    df_ly = df[df["APS"] == derniere_ly].groupby("Code_SAQ").agg(
        Quantite_LY=("Quantite", "sum"),
        Montant_LY =("Montant",  "sum"),
    ).reset_index()

    # ── Cumul été 2026 (toutes semaines disponibles) ──────────────────
    df_cum26 = df[df["Annee"] == 2026].groupby("Code_SAQ").agg(
        Cumul_Ete_2026_Btl=("Quantite", "sum"),
        Cumul_Ete_2026    =("Montant",  "sum"),
    ).reset_index()

    # ── Cumul été 2025 — MÊMES semaines que 2026 en cours ────────────
    # Convertir chaque semaine 2026 en équivalent 2025
    aps_2025_equiv = []
    for aps in semaines_2026:
        p = aps.split("-")
        aps_2025_equiv.append(f"{int(p[0])-1}-{p[1]}-{p[2]}")

    df_cum25 = df[df["APS"].isin(aps_2025_equiv)].groupby("Code_SAQ").agg(
        Cumul_Ete_2025_Btl=("Quantite", "sum"),
        Cumul_Ete_2025    =("Montant",  "sum"),
    ).reset_index()

    # ── Merge ─────────────────────────────────────────────────────────
    df_s = df_s.merge(df_ly,    on="Code_SAQ", how="left")
    df_s = df_s.merge(df_cum26, on="Code_SAQ", how="left")
    df_s = df_s.merge(df_cum25, on="Code_SAQ", how="left")

    for col in ["Quantite_LY","Montant_LY",
                "Cumul_Ete_2026","Cumul_Ete_2026_Btl",
                "Cumul_Ete_2025","Cumul_Ete_2025_Btl"]:
        df_s[col] = df_s[col].fillna(0)

    # Variations
    df_s["Var_Btl_pct"] = df_s.apply(
        lambda r: (r["Quantite"] - r["Quantite_LY"]) / r["Quantite_LY"] * 100
        if r["Quantite_LY"] > 0 else None, axis=1)
    df_s["Var_pct"] = df_s.apply(
        lambda r: (r["Montant"] - r["Montant_LY"]) / r["Montant_LY"] * 100
        if r["Montant_LY"] > 0 else None, axis=1)
    df_s["Var_Cum_pct"] = df_s.apply(
        lambda r: (r["Cumul_Ete_2026"] - r["Cumul_Ete_2025"]) / r["Cumul_Ete_2025"] * 100
        if r["Cumul_Ete_2025"] > 0 else None, axis=1)

    df_s = df_s.sort_values("Montant", ascending=False).reset_index(drop=True)
    return df_s, derniere, semaines_2026

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 : TOUS LES PRODUITS
# ══════════════════════════════════════════════════════════════════════════════

def onglet_produits(wb, df, semaine, semaines_2026):
    ws  = wb.create_sheet("Ventes Semaine")
    nb_sem = len(semaines_2026)
    titre_page(ws, f"VENTES PAB — Semaine {semaine}  |  Cumul : {semaines_2026[0]} → {semaines_2026[-1]}", BLEU_FONCE, 13)

    row = 3
    headers = [
        "Rang", "Code SAQ", "Description", "Marque", "Segment",
        "Btl Semaine", "Var. Btl",
        "Ventes Semaine $", "Var. $",
        "Btl Été 2026", "Btl Été 2025",
        "Cumul Été 2026 $", "Cumul Été 2025 $", "Var. Cumul"
    ]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=col, value=h))
    row += 1

    for rang, (_, r) in enumerate(df.iterrows(), 1):
        is_lab = r.get("Marque_PaB","") in MARQUES_LABATT
        bg = BLEU_CLAIR if is_lab else None
        vals = [
            rang,
            r["Code_SAQ"],
            r.get("Description",""),
            r.get("Marque_PaB",""),
            r.get("Segment",""),
            int(r["Quantite"]),
            pct_str(r.get("Var_Btl_pct")),
            round(r["Montant"], 2),
            pct_str(r.get("Var_pct")),
            int(r.get("Cumul_Ete_2026_Btl", 0)),
            int(r.get("Cumul_Ete_2025_Btl", 0)),
            round(r.get("Cumul_Ete_2026", 0), 2),
            round(r.get("Cumul_Ete_2025", 0), 2),
            pct_str(r.get("Var_Cum_pct")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=is_lab, bg=bg,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [6, 10, 11]: cell.number_format = FMT_NOMBRE
            if col in [8, 12, 13]: cell.number_format = FMT_DOLLAR
        row += 1

    # Total
    tot = ["", "TOTAL", "", "", "",
           int(df["Quantite"].sum()), "",
           round(df["Montant"].sum(), 2), "",
           int(df["Cumul_Ete_2026_Btl"].sum()),
           int(df["Cumul_Ete_2025_Btl"].sum()),
           round(df["Cumul_Ete_2026"].sum(), 2),
           round(df["Cumul_Ete_2025"].sum(), 2), ""]
    for col, val in enumerate(tot, 1):
        cell = ws.cell(row=row, column=col, value=val)
        style_total(cell)
        add_border(cell)
        if col in [6, 10, 11]: cell.number_format = FMT_NOMBRE
        if col in [8, 12, 13]: cell.number_format = FMT_DOLLAR

    auto_width(ws)
    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 30

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 : INNOVATIONS
# ══════════════════════════════════════════════════════════════════════════════

def onglet_innovations(wb, df, semaine):
    df_inno = df[
        (pd.to_numeric(df["Code_SAQ"], errors="coerce") >= SEUIL_INNO) |
        (df["Montant_LY"] == 0)
    ].sort_values("Montant", ascending=False).reset_index(drop=True)

    ws = wb.create_sheet("Innovations")
    titre_page(ws, f"INNOVATIONS PAB — Semaine {semaine}  ({len(df_inno)} produits)", VERT_FONCE, 10)

    row = 3
    headers = ["Rang Global", "Code SAQ", "Description", "Marque", "Segment",
               "Btl Semaine", "Ventes Semaine $",
               "Cumul Été 2026 Btl", "Cumul Été 2026 $",
               "Présent 2025?"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=VERT_FONCE)
    row += 1

    for _, r in df_inno.iterrows():
        rang_global = df[df["Code_SAQ"] == r["Code_SAQ"]].index[0] + 1
        is_lab = r.get("Marque_PaB","") in MARQUES_LABATT
        bg = BLEU_CLAIR if is_lab else VERT_CLAIR
        present = "Non" if r["Montant_LY"] == 0 else "Oui"
        vals = [rang_global, r["Code_SAQ"], r.get("Description",""),
                r.get("Marque_PaB",""), r.get("Segment",""),
                int(r["Quantite"]), round(r["Montant"], 2),
                int(r.get("Cumul_Ete_2026_Btl", 0)),
                round(r.get("Cumul_Ete_2026", 0), 2), present]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=is_lab, bg=bg,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [6, 8]: cell.number_format = FMT_NOMBRE
            if col in [7, 9]: cell.number_format = FMT_DOLLAR
        row += 1

    auto_width(ws)
    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 : MARQUES
# ══════════════════════════════════════════════════════════════════════════════

def onglet_marques(wb, df, semaine):
    ws = wb.create_sheet("Marques")
    titre_page(ws, f"VENTES PAR MARQUE PAB — Semaine {semaine}", ORANGE_FONCE, 11)

    row = 3
    headers = ["Rang", "Marque", "Nb Produits",
               "Btl Semaine", "Btl LY", "Var. Btl",
               "Ventes Semaine $", "Ventes LY $", "Var. $",
               "Cumul Été 2026 $", "Cumul Été 2025 $", "Var. Cumul"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=ORANGE_FONCE)
    row += 1

    grp = df.groupby("Marque_PaB").agg(
        Nb_produits       =("Code_SAQ",           "nunique"),
        Btl               =("Quantite",            "sum"),
        Btl_LY            =("Quantite_LY",         "sum"),
        Amt               =("Montant",             "sum"),
        Amt_LY            =("Montant_LY",          "sum"),
        Cum26             =("Cumul_Ete_2026",       "sum"),
        Cum25             =("Cumul_Ete_2025",       "sum"),
        Cum26_btl         =("Cumul_Ete_2026_Btl",  "sum"),
        Cum25_btl         =("Cumul_Ete_2025_Btl",  "sum"),
    ).reset_index().sort_values("Amt", ascending=False).reset_index(drop=True)

    total_amt = grp["Amt"].sum()

    for rang, (_, r) in enumerate(grp.iterrows(), 1):
        vb   = (r["Btl"] - r["Btl_LY"]) / r["Btl_LY"] * 100 if r["Btl_LY"] > 0 else None
        va   = (r["Amt"] - r["Amt_LY"]) / r["Amt_LY"] * 100 if r["Amt_LY"] > 0 else None
        vc   = (r["Cum26"] - r["Cum25"]) / r["Cum25"] * 100 if r["Cum25"] > 0 else None
        is_lab = r["Marque_PaB"] in MARQUES_LABATT
        bg = BLEU_CLAIR if is_lab else ORANGE_CLAIR

        vals = [rang, r["Marque_PaB"], int(r["Nb_produits"]),
                int(r["Btl"]), int(r["Btl_LY"]), pct_str(vb),
                round(r["Amt"],2), round(r["Amt_LY"],2), pct_str(va),
                round(r["Cum26"],2), round(r["Cum25"],2), pct_str(vc)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=is_lab, bg=bg,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [4, 5]: cell.number_format = FMT_NOMBRE
            if col in [7, 8, 10, 11]: cell.number_format = FMT_DOLLAR
        row += 1

    # Total
    tot_vb = (grp["Btl"].sum() - grp["Btl_LY"].sum()) / grp["Btl_LY"].sum() * 100 if grp["Btl_LY"].sum() > 0 else None
    tot_va = (grp["Amt"].sum() - grp["Amt_LY"].sum()) / grp["Amt_LY"].sum() * 100 if grp["Amt_LY"].sum() > 0 else None
    tot_vc = (grp["Cum26"].sum() - grp["Cum25"].sum()) / grp["Cum25"].sum() * 100 if grp["Cum25"].sum() > 0 else None
    tot = ["", "TOTAL", int(grp["Nb_produits"].sum()),
           int(grp["Btl"].sum()), int(grp["Btl_LY"].sum()), pct_str(tot_vb),
           round(grp["Amt"].sum(),2), round(grp["Amt_LY"].sum(),2), pct_str(tot_va),
           round(grp["Cum26"].sum(),2), round(grp["Cum25"].sum(),2), pct_str(tot_vc)]
    for col, val in enumerate(tot, 1):
        cell = ws.cell(row=row, column=col, value=val)
        style_total(cell)
        add_border(cell)
        if col in [4, 5]: cell.number_format = FMT_NOMBRE
        if col in [7, 8, 10, 11]: cell.number_format = FMT_DOLLAR

    auto_width(ws)
    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 : SEGMENTS
# ══════════════════════════════════════════════════════════════════════════════

def onglet_segments(wb, df, semaine):
    ws = wb.create_sheet("Segments")
    titre_page(ws, f"VENTES PAR SEGMENT PAB — Semaine {semaine}", VIOLET_FONCE, 11)

    row = 3
    headers = ["Rang", "Segment", "Nb Produits",
               "Btl Semaine", "Btl LY", "Var. Btl",
               "Ventes Semaine $", "Ventes LY $", "Var. $",
               "Cumul Été 2026 $", "Cumul Été 2025 $", "Var. Cumul"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=VIOLET_FONCE)
    row += 1

    grp = df[df["Segment"] != ""].groupby("Segment").agg(
        Nb_produits=("Code_SAQ",          "nunique"),
        Btl        =("Quantite",           "sum"),
        Btl_LY     =("Quantite_LY",        "sum"),
        Amt        =("Montant",            "sum"),
        Amt_LY     =("Montant_LY",         "sum"),
        Cum26      =("Cumul_Ete_2026",      "sum"),
        Cum25      =("Cumul_Ete_2025",      "sum"),
    ).reset_index().sort_values("Amt", ascending=False).reset_index(drop=True)

    for rang, (_, r) in enumerate(grp.iterrows(), 1):
        vb = (r["Btl"] - r["Btl_LY"]) / r["Btl_LY"] * 100 if r["Btl_LY"] > 0 else None
        va = (r["Amt"] - r["Amt_LY"]) / r["Amt_LY"] * 100 if r["Amt_LY"] > 0 else None
        vc = (r["Cum26"] - r["Cum25"]) / r["Cum25"] * 100 if r["Cum25"] > 0 else None
        vals = [rang, r["Segment"], int(r["Nb_produits"]),
                int(r["Btl"]), int(r["Btl_LY"]), pct_str(vb),
                round(r["Amt"],2), round(r["Amt_LY"],2), pct_str(va),
                round(r["Cum26"],2), round(r["Cum25"],2), pct_str(vc)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bg=VIOLET_CLAIR,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [4, 5]: cell.number_format = FMT_NOMBRE
            if col in [7, 8, 10, 11]: cell.number_format = FMT_DOLLAR
        row += 1

    # Total
    tot_vb = (grp["Btl"].sum() - grp["Btl_LY"].sum()) / grp["Btl_LY"].sum() * 100 if grp["Btl_LY"].sum() > 0 else None
    tot_va = (grp["Amt"].sum() - grp["Amt_LY"].sum()) / grp["Amt_LY"].sum() * 100 if grp["Amt_LY"].sum() > 0 else None
    tot_vc = (grp["Cum26"].sum() - grp["Cum25"].sum()) / grp["Cum25"].sum() * 100 if grp["Cum25"].sum() > 0 else None
    tot = ["", "TOTAL", int(grp["Nb_produits"].sum()),
           int(grp["Btl"].sum()), int(grp["Btl_LY"].sum()), pct_str(tot_vb),
           round(grp["Amt"].sum(),2), round(grp["Amt_LY"].sum(),2), pct_str(tot_va),
           round(grp["Cum26"].sum(),2), round(grp["Cum25"].sum(),2), pct_str(tot_vc)]
    for col, val in enumerate(tot, 1):
        cell = ws.cell(row=row, column=col, value=val)
        style_total(cell)
        add_border(cell)
        if col in [4, 5]: cell.number_format = FMT_NOMBRE
        if col in [7, 8, 10, 11]: cell.number_format = FMT_DOLLAR

    auto_width(ws)
    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 5 : PARTS DE MARCHÉ
# ══════════════════════════════════════════════════════════════════════════════

def onglet_pm(wb, df, semaine, semaines_2026):
    ws = wb.create_sheet("Parts de Marche")
    titre_page(ws, f"PARTS DE MARCHÉ LABATT PAB — Semaine {semaine}", BLEU_FONCE, 8)

    row = 3
    df_lab = df[df["Marque_PaB"].isin(MARQUES_LABATT)]

    total_btl    = df["Quantite"].sum()
    total_btl_ly = df["Quantite_LY"].sum()
    total_amt    = df["Montant"].sum()
    total_amt_ly = df["Montant_LY"].sum()
    lab_btl      = df_lab["Quantite"].sum()
    lab_btl_ly   = df_lab["Quantite_LY"].sum()
    lab_amt      = df_lab["Montant"].sum()
    lab_amt_ly   = df_lab["Montant_LY"].sum()
    pm_sem       = lab_amt / total_amt * 100 if total_amt > 0 else 0

    cum26_amt    = df["Cumul_Ete_2026"].sum()
    cum25_amt    = df["Cumul_Ete_2025"].sum()
    cum26_btl    = df["Cumul_Ete_2026_Btl"].sum()
    cum25_btl    = df["Cumul_Ete_2025_Btl"].sum()
    lab_c26_amt  = df_lab["Cumul_Ete_2026"].sum()
    lab_c25_amt  = df_lab["Cumul_Ete_2025"].sum()
    lab_c26_btl  = df_lab["Cumul_Ete_2026_Btl"].sum()
    lab_c25_btl  = df_lab["Cumul_Ete_2025_Btl"].sum()
    pm_cum       = lab_c26_amt / cum26_amt * 100 if cum26_amt > 0 else 0

    # ── Semaine ───────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="SEMAINE EN COURS").font = Font(bold=True, size=11, color=GRIS_HEADER, name="Arial")
    row += 1

    for col, h in enumerate(["", "Btl 2026", "Btl 2025", "Var. Btl", "$ 2026", "$ 2025", "Var. $", "PM Labatt"], 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=GRIS_HEADER)
    row += 1

    for label, btl, btl_ly, amt, amt_ly in [
        ("Total PAB",  total_btl, total_btl_ly, total_amt, total_amt_ly),
        ("Labatt PAB", lab_btl,   lab_btl_ly,   lab_amt,   lab_amt_ly),
    ]:
        vb  = (btl - btl_ly) / btl_ly * 100 if btl_ly > 0 else None
        va  = (amt - amt_ly) / amt_ly * 100 if amt_ly > 0 else None
        pm  = f"{pm_sem:.1f}%" if label == "Labatt PAB" else ""
        vals = [label, int(btl), int(btl_ly), pct_str(vb), round(amt,2), round(amt_ly,2), pct_str(va), pm]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=True, bg=BLEU_CLAIR,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [2, 3]: cell.number_format = FMT_NOMBRE
            if col in [5, 6]: cell.number_format = FMT_DOLLAR
        row += 1

    row += 1

    # ── Cumul été ─────────────────────────────────────────────────────
    nb_sem = len(semaines_2026)
    ws.cell(row=row, column=1, value=f"CUMUL ÉTÉ ({semaines_2026[0]} → {semaines_2026[-1]})").font = Font(bold=True, size=11, color=GRIS_HEADER, name="Arial")
    row += 1

    for col, h in enumerate(["", "Btl 2026", "Btl 2025", "Var. Btl", "$ 2026", "$ 2025", "Var. $", "PM Labatt"], 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=GRIS_HEADER)
    row += 1

    for label, btl26, btl25, amt26, amt25 in [
        ("Total PAB",  cum26_btl, cum25_btl, cum26_amt, cum25_amt),
        ("Labatt PAB", lab_c26_btl, lab_c25_btl, lab_c26_amt, lab_c25_amt),
    ]:
        vb  = (btl26 - btl25) / btl25 * 100 if btl25 > 0 else None
        va  = (amt26 - amt25) / amt25 * 100 if amt25 > 0 else None
        pm  = f"{pm_cum:.1f}%" if label == "Labatt PAB" else ""
        vals = [label, int(btl26), int(btl25), pct_str(vb), round(amt26,2), round(amt25,2), pct_str(va), pm]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=True, bg=BLEU_CLAIR,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [2, 3]: cell.number_format = FMT_NOMBRE
            if col in [5, 6]: cell.number_format = FMT_DOLLAR
        row += 1

    row += 2

    # ── Détail par marque Labatt ──────────────────────────────────────
    ws.cell(row=row, column=1, value="DÉTAIL PAR MARQUE LABATT").font = Font(bold=True, size=11, color=GRIS_HEADER, name="Arial")
    row += 1

    for col, h in enumerate(["Marque","Btl Sem.","Btl LY","Var. Btl","$ Sem.","$ LY","Var. $","PM Sem.","Cum. 2026","Cum. 2025","Var. Cum."], 1):
        style_header(ws.cell(row=row, column=col, value=h))
    row += 1

    pm_marques = df_lab.groupby("Marque_PaB").agg(
        Btl      =("Quantite",           "sum"),
        Btl_LY   =("Quantite_LY",        "sum"),
        Amt      =("Montant",            "sum"),
        Amt_LY   =("Montant_LY",         "sum"),
        Cum26    =("Cumul_Ete_2026",      "sum"),
        Cum25    =("Cumul_Ete_2025",      "sum"),
    ).reset_index().sort_values("Amt", ascending=False)

    for _, r in pm_marques.iterrows():
        vb  = (r["Btl"] - r["Btl_LY"]) / r["Btl_LY"] * 100 if r["Btl_LY"] > 0 else None
        va  = (r["Amt"] - r["Amt_LY"]) / r["Amt_LY"] * 100 if r["Amt_LY"] > 0 else None
        vc  = (r["Cum26"] - r["Cum25"]) / r["Cum25"] * 100 if r["Cum25"] > 0 else None
        pm_m = r["Amt"] / total_amt * 100 if total_amt > 0 else 0
        vals = [r["Marque_PaB"], int(r["Btl"]), int(r["Btl_LY"]), pct_str(vb),
                round(r["Amt"],2), round(r["Amt_LY"],2), pct_str(va),
                f"{pm_m:.1f}%", round(r["Cum26"],2), round(r["Cum25"],2), pct_str(vc)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bold=True, bg=BLEU_CLAIR,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col in [2, 3]: cell.number_format = FMT_NOMBRE
            if col in [5, 6, 9, 10]: cell.number_format = FMT_DOLLAR
        row += 1

    auto_width(ws)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 6 : À CLASSIFIER
# ══════════════════════════════════════════════════════════════════════════════

def onglet_classifier(wb, df, semaine):
    df_cl = df[
        df["Marque_PaB"].isna() |
        (df["Marque_PaB"] == "") |
        (df["Marque_PaB"] == "nan")
    ].sort_values("Montant", ascending=False)

    ws = wb.create_sheet("A Classifier")
    titre_page(ws, f"À CLASSIFIER — Semaine {semaine}  ({len(df_cl)} produits)", ROUGE_FONCE, 7)

    row = 3
    if len(df_cl) == 0:
        ws.cell(row=row, column=1, value="✅ Aucun produit à classifier cette semaine !")
        ws.cell(row=row, column=1).font = Font(bold=True, color=VERT_FONCE, name="Arial")
        return

    for col, h in enumerate(["Rang","Code SAQ","Description","Segment","Btl Semaine","Ventes $ Semaine","Cumul Été 2026 $"], 1):
        style_header(ws.cell(row=row, column=col, value=h), bg=ROUGE_FONCE)
    row += 1

    for _, r in df_cl.iterrows():
        rang_global = df[df["Code_SAQ"] == r["Code_SAQ"]].index[0] + 1
        vals = [rang_global, r["Code_SAQ"], r.get("Description",""),
                r.get("Segment",""), int(r["Quantite"]),
                round(r["Montant"],2), round(r.get("Cumul_Ete_2026",0),2)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, bg=ROUGE_CLAIR,
                       align="right" if isinstance(val, (int, float)) else "left")
            add_border(cell)
            if col == 5: cell.number_format = FMT_NOMBRE
            if col in [6, 7]: cell.number_format = FMT_DOLLAR
        row += 1

    auto_width(ws)
    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═"*60)
    print("  RAPPORT HEBDOMADAIRE PAB — Labatt  v3")
    print("═"*60)

    print("\n1. Chargement du CSV...")
    df_raw = charger_donnees()

    print("\n2. Préparation des données...")
    df, semaine, semaines_2026 = preparer(df_raw)

    df_inno = df[(pd.to_numeric(df["Code_SAQ"], errors="coerce") >= SEUIL_INNO) | (df["Montant_LY"] == 0)]
    df_cl   = df[df["Marque_PaB"].isna() | (df["Marque_PaB"] == "") | (df["Marque_PaB"] == "nan")]

    print(f"   ✅ {len(df)} produits PAB")
    print(f"   ✅ {len(df_inno)} innovations")
    print(f"   ✅ {df['Marque_PaB'].nunique()} marques")
    print(f"   ✅ {df['Segment'].nunique()} segments")
    print(f"   ⚠️  {len(df_cl)} à classifier")
    print(f"   ✅ Cumul basé sur {len(semaines_2026)} semaines 2026")

    print("\n3. Génération du rapport Excel...")
    wb = Workbook()
    wb.remove(wb.active)
    onglet_produits(wb, df, semaine, semaines_2026)
    onglet_innovations(wb, df, semaine)
    onglet_marques(wb, df, semaine)
    onglet_segments(wb, df, semaine)
    onglet_pm(wb, df, semaine, semaines_2026)
    onglet_classifier(wb, df, semaine)

    output = f"Rapport_PAB_{semaine}.xlsx"
    wb.save(output)
    print(f"\n   ✅ Rapport sauvegardé : {output}")
    print("   📋 Onglets : Ventes Semaine | Innovations | Marques | Segments | Parts de Marche | A Classifier")
    print("\n" + "═"*60 + "\n")

if __name__ == "__main__":
    main()