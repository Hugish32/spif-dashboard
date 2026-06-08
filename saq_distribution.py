import requests
import re
import time
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PRODUITS = [
    ("11927661", "Smirnoff Ice Canette (6x355 ml)"),
    ("14937996", "Mott's Clamato Caesar Original 6 x 341 ml"),
    ("15298102", "Disaronno Sour (4 x 355 ml)"),
    ("14699633", "Smirnoff Vodka et Soda Rosee 12 x 355 ML"),
    ("13991835", "Seventh Heaven Gin Tonic 4 x 355 ml"),
    ("15298807", "Mott's Clamato Caesar Extra- epice (6 x 341 ml)"),
    ("13990912", "Smirnoff Ice Berry (6x355 ml) Cooler"),
    ("542977", "Rev Guarana & Ginseng (4 x 330 ml)"),
    ("15455340", "Pur Vodka Cosmo Canneberge Grenade et Agrumes 4 X 355 ml"),
    ("15298321", "Captain Morgan Collection de Cocktails (12 X 355ml)"),
    ("14698016", "3 Lacs Gin Lim Rose Pamplemous Romarin (4 X 355ML)"),
    ("14536367", "White Claw Caisse Mixte No.1 (12 X 355 ML)"),
    ("15585814", "Popsicle Firecraker Vodka Cocktail (4x355ml)"),
    ("15455585", "3 Lacs Gin Limonade Peche Passion 4 x 355ml"),
    ("15313499", "Romeo's Gin + Tonic (4x355 ml)"),
    ("11927670", "Smirnoff Ice Bouteille (24 x 330 ml)"),
    ("15299463", "3 Lacs Limonade Framboise Bleue (4 x 355ml)"),
    ("15148358", "Captain Morgan Mai Tai Mangue (4 x 355 ml)"),
    ("14699035", "Miele Sour Limonade a l'Amaretto (4 x 355 ml)"),
    ("15372093", "Mott's Clamato Caesar Haricot Marine"),
    ("14397993", "Seventh Heaven Fraise Sauvage et Rhubarb Gin Fizz 4 x 355 ml"),
    ("14969980", "Smirnoff Ice du Party Plein la Caisse (12X355ml)"),
    ("15570831", "Smirnoff Ice Light Variety Pack (12 x 355 ml)"),
    ("15557119", "White Claw Caisse Mixte No. 2 12 x 355 ml"),
    ("15570794", "3 Lacs Gin Limonade Cerise Lime (4 x 355 ml)"),
    ("15565696", "Mott's Clamato Caesar Original"),
    ("15572714", "Romeo's Gin Clementine Limonade (4 x 355 ml)"),
    ("15372122", "Mott's Clamato Caesar Le Tout Garni"),
    ("15570823", "White Claw Clawtails (12 x 355 ml)"),
    ("15580191", "3 Lacs Gin Limonade Rose (3L)"),
    ("13612137", "Lolea Numero Un"),
    ("15162221", "Palm Bay Frozen Rainbow Twist"),
    ("15455382", "Smirnoff Cocktails Collection 12 x 355 ml"),
    ("15570807", "Seventh Heaven Fleur de Sureau et Poire (4 x 355 ml)"),
    ("13248646", "Black Fly Vodka 4x400 ml Grapefruit"),
    ("15585689", "Popsicle Vodka Soda Caisse Mixte (12x355ml)"),
    ("15544570", "Beach Day Every Day Vodka Frozen Berries 4 x 355 ml"),
    ("15455622", "Captain Morgan Guava Mojito 4 x 355 ml"),
    ("11927741", "Smirnoff Ice (4 x 330 ml)"),
    ("15570778", "Captain Morgan Spiced Rum Punch (12 x 355 ml)"),
    ("10933998", "Mudshake Vanille Francaise (4x270 ml)"),
    ("15103598", "White Claw Hard Seltzer Surge Caisse Mixte (12 X 355ml)"),
    ("15646541", "Bacardi Mojito (4x355ml)"),
    ("15298330", "Smirnoff Vodka & Soda Caisse Mixte (12 x 355ml)"),
    ("15570760", "White Claw Big Surf (12 x 355 ml)"),
    ("15148905", "Mott's Clamato Caesar Saveurs Assorties (12 x 341 ml)"),
    ("15585726", "Black Fly Blue Lagoon (4x400ml)"),
    ("15455366", "Ward Cola Diete 4 x 355 ml"),
    ("15585865", "Pinata Peche Surette (4x355ml)"),
    ("14699617", "Chic Choc Rhum Epice Libre (4 x 355 ml)"),
    ("14959562", "Smirnoff Seltzer Limonades (12 X 355ml)"),
    ("15472756", "Jameson Limonade au Whiskey 4 x 355 ml"),
    ("15381926", "Happy Dad Hard Seltzer Caisse Assortie (12X355 ml)"),
    ("14701334", "Twisted Tea Originale"),
    ("15555121", "Coors Seltzer Slushie Variety Pack (12 x 355ml)"),
    ("15298090", "Malibu Saveurs Assorties (8 x 355 ml)"),
    ("15570858", "Bacardi Limon Punch (12 x 355 ml)"),
    ("15148016", "Palm Bay Fraise Kiwi (4 x 355 ml)"),
    ("14700235", "Smirnoff Ice Framboise (4 x 330 ml)"),
    ("14398961", "Palm Bay Peche Mangue (4 x 355 ml)"),
    ("15148032", "Palm Bay Framboise Bleue Citron (4 x 355 ml)"),
    ("15299586", "Twisted Tea Peche"),
    ("14990949", "Kopparberg Fraise et Lime (4 x 500 ml)"),
    ("15585515", "Pinata Fraise Acide (4x355ml)"),
    ("13248638", "Black Fly Cranberry Cooler (4x400 ml)"),
    ("15585478", "Pinata Framboise Bleue (4x355ml)"),
    ("15588935", "Bacardi Punch Tropical (4x355ml)"),
    ("15372114", "Mott's Clamato Caesar Cornichon"),
    ("15455307", "Smirnoff Seltzer Rose (12 x 355ml)"),
    ("15404104", "Caisse Mixte Mott's Clamato Caesar (12 x 341 ml)"),
    ("14699561", "Chic Choc Rhum Epice et Gingembre (4 x 355 ml)"),
    ("15298532", "Smirnoff Seltzer Citron (4 x 355 ml)"),
    ("15298479", "Smirnoff Seltzer Framboise (4 x 355 ml)"),
    ("15570719", "Smirnoff Ice Tropical (12 x 355 ml)"),
    ("15646550", "Bacardi Limon Punch (4x355ml)"),
    ("15455391", "Smirnoff Seltzer Fraise (4 x 355 ml)"),
    ("15585777", "Pinata Mangue Piquante (4x355ml)"),
    ("15570743", "Captain Morgan Cocktail Tropical Punch (4 x 355 ml)"),
    ("15474217", "Carabins Gin Tonic (4 x 355 ml)"),
    ("14961734", "Kopparberg Poire (4 x 500 ml)"),
    ("15455411", "Smirnoff Seltzer Peche (4 x 355 ml)"),
    ("14959693", "Smirnoff Seltzer Framboise Sauvage (4 x 355 ml)"),
    ("15611584", "Twisted Tea Demi-Demi"),
    ("15296756", "Caisse Mixte Twisted Tea"),
    ("15372106", "Mott's Clamato Caesar Lime"),
    ("15455518", "Carabins Gin Framboise (4 x 355 ml)"),
    ("15298372", "Smirnoff Seltzer Citron Vert (4 x 355 ml)"),
    ("14398777", "Palm Bay Citron Lime (4 x 355 ml)"),
    ("15455551", "Carabins Gin Peche (4 x 355 ml)"),
    ("15599431", "Mott's Clamato Caesar Sriracha"),
    ("14699596", "Chic Choc Rhum Epice et Cola (4 x 355 ml)"),
    ("15298081", "Malibu Ananas (4 x 355 ml)"),
    ("15455462", "Carabins Gin Citron (4 x 355 ml)"),
    ("15162117", "Palm Bay Ananas Citron Vert (4 x 355 ml)"),
    ("15585873", "Pinata Citron Vert Piquant (4x355ml)"),
    ("15148382", "Palm Bay Mangue Passion (4 x 355 ml)"),
    ("15148649", "Palm Bay Fraise Citron (4 x 355 ml)"),
    ("14700497", "Smirnoff Ice Fraise (4 x 330 ml)"),
    ("15601589", "Bacardi Punch Tropical (12x355ml)"),
    ("15586921", "Twisted Tea Citron"),
    ("15570735", "White Claw Variety Pack No.3 (12 x 355 ml)"),
    ("15585460", "Pinata Ananas Noix de Coco (4x355ml)"),
    ("15550419", "Carabins Caisse Mixte (12 x 355 ml)"),
    ("14398751", "Palm Bay Framboise Peche (4 x 355 ml)"),
    ("15149094", "Palm Bay Citron Lime Zero Sucre (4 x 355 ml)"),
    ("15585742", "Pinata Melon d'Eau Citron Vert (4x355ml)"),
    ("15301261", "Caisse Mixte Carabins (8 x 355 ml)"),
    ("15148075", "Palm Bay Peche Citron (4 x 355 ml)"),
    ("15455374", "Smirnoff Seltzer Mangue (4 x 355 ml)"),
    ("15455657", "Carabins Gin Bleuet (4 x 355 ml)"),
    ("15455358", "Smirnoff Seltzer Grenadine (4 x 355 ml)"),
    ("15585911", "Pinata Fraise Kiwi (4x355ml)"),
    ("15547607", "Carabins Gin Limonade Rose (4 x 355 ml)"),
    ("15148121", "Palm Bay Fraise Ananas (4 x 355 ml)"),
    ("15298014", "Malibu Cola (4 x 355 ml)"),
    ("15585857", "Pinata Cerises Noires (4x355ml)"),
    ("15608342", "Twisted Tea Citron Limonade"),
    ("14016020", "Mott's Clamato Caesar Classique"),
    ("15455294", "Smirnoff Seltzer Citron Vert (12 x 355ml)"),
    ("14968135", "Kopparberg Fruits des Bois (4 x 500 ml)"),
    ("15570840", "Malibu Caisse Mixte (12 x 355 ml)"),
    ("15299914", "Mike's Hard Lemonade"),
    ("15455631", "Captain Morgan Cocktail Citron Vert Mojito (4 x 355 ml)"),
    ("15313481", "Romeo's Gin Framboise Limonade (4x355 ml)"),
    ("15298006", "Malibu Ananas Zero (4 x 355 ml)"),
    ("15455403", "Smirnoff Seltzer Framboise (12 x 355ml)"),
    ("15585582", "Pinata Citrus (4x355ml)"),
    ("15586981", "Twisted Tea Framboise"),
    ("15570903", "Smirnoff Ice Limonade (12 x 355 ml)"),
    ("15299412", "Mike's Hard Lemonade Rose"),
    ("15588353", "Bacardi Limon Punch Tropical (4x355ml)"),
    ("15455526", "Carabins Gin Limonade (4 x 355 ml)"),
    ("15305579", "Caisse Mixte Carabins (12 x 355 ml)"),
    ("15570727", "Smirnoff Ice Framboise (12 x 355 ml)"),
    ("15528730", "Carabins Gin Melon (4 x 355 ml)"),
    ("15586956", "Twisted Tea Peche Mangue"),
    ("15570751", "Smirnoff Ice Tropical (4 x 355 ml)"),
    ("15455497", "Carabins Gin Fraise (4 x 355 ml)"),
    ("15029412", "Kopparberg Citron (4 x 500 ml)"),
    ("15637282", "Mate Libre Mojito (4 x 355 ml)"),
    ("14883553", "Mate Libre Eau Petillante Citron (4 x 355 ml)"),
    ("15587019", "Twisted Tea Mangue"),
    ("15614857", "Twisted Tea Fraise"),
    ("15585697", "Popsicle Firecraker Vodka Cocktail Caisse Mixte (12x355ml)"),
    ("15455446", "Carabins Gin Canneberge (4 x 355 ml)"),
    ("15585970", "Pinata Melon d'Eau (4x355ml)"),
    ("15549389", "Carabins Gin Limonade Bleue (4 x 355 ml)"),
    ("15471534", "Carabins Gin Ananas (4 x 355 ml)"),
    ("15455534", "Carabins Gin Citron Vert (4 x 355 ml)"),
    ("14559841", "Smirnoff Ice Tropical (4 x 355 ml)"),
    ("14959773", "Smirnoff Seltzer Citron (12 x 355ml)"),
    ("14300282", "Mott's Clamato Caesar Epice"),
    ("15455438", "Carabins Gin Fraise Kiwi (4 x 355 ml)"),
    ("15299949", "Mike's Hard Cranberry Lemonade"),
    ("15518321", "Carabins Gin Framboise Bleue (4 x 355 ml)"),
    ("15020442", "Kopparberg Fraise (4 x 500 ml)"),
    ("14700121", "Smirnoff Ice Citron (4 x 330 ml)"),
    ("15077895", "Kopparberg Peche et Framb (4 x 500 ml)"),
    ("15462793", "Carabins Gin Mure (4 x 355 ml)"),
    ("15149238", "Palm Bay Framboise Peche Zero Sucre (4 x 355 ml)"),
    ("15298356", "Smirnoff Seltzer Fraise (12 x 355ml)"),
    ("15586948", "Twisted Tea Originale Bouteille"),
    ("15455649", "Carabins Gin Kiwi (4 x 355 ml)"),
    ("15149203", "Palm Bay Fraise Kiwi Zero Sucre (4 x 355 ml)"),
    ("15626292", "Twisted Tea Limonade"),
    ("15148307", "Palm Bay Canneberge Citron (4 x 355 ml)"),
    ("15585620", "Popsicle Strawberry Shortcake Vodka Cocktail (4x355ml)"),
    ("14700075", "Smirnoff Ice Melon (4 x 330 ml)"),
    ("15528625", "Carabins Gin Concombre (4 x 355 ml)"),
    ("15525301", "Carabins Gin Pamplemousse (4 x 355 ml)"),
    ("14699588", "Chic Choc Rhum Epice et Gingembre Zero Sucre (4 x 355 ml)"),
    ("15455331", "Smirnoff Seltzer Mangue (12 x 355ml)"),
    ("15586999", "Twisted Tea Demi-Demi Bouteille"),
    ("15498227", "Carabins Gin Pomme Verte (4 x 355 ml)"),
    ("15168156", "Carabins Gin Citron Vert Zero (4 x 355 ml)"),
    ("15545185", "Carabins Gin Limonade Fraise (4 x 355 ml)"),
    ("15311485", "Romeo's Gin Tonic Caisse Mixte (8 x 355 ml)"),
    ("15455454", "Carabins Gin Limonade Citron (4 x 355 ml)"),
    ("15574277", "Carabins Gin Limonade Framboise (4 x 355 ml)"),
    ("15126976", "Kopparberg Mangue et Framb (4 x 500 ml)"),
    ("15455315", "Smirnoff Seltzer Peche (12 x 355ml)"),
    ("15574162", "Carabins Gin Limonade Peche (4 x 355 ml)"),
    ("14959491", "Smirnoff Seltzer Framboise Sauvage (12 x 355ml)"),
    ("15498219", "Carabins Gin Cerises (4 x 355 ml)"),
    ("15148139", "Palm Bay Framboise Citron (4 x 355 ml)"),
    ("14889648", "Mate Libre Eau Petillante Orange (4 x 355 ml)"),
    ("15455471", "Carabins Gin Tropical (4 x 355 ml)"),
    ("15055709", "Kopparberg Citron et Lime (4 x 500 ml)"),
    ("15298313", "Smirnoff Seltzer Grenadine (12 x 355ml)"),
    ("15148235", "Palm Bay Fraise Citron Zero Sucre (4 x 355 ml)"),
    ("15298188", "Malibu Fraise (4 x 355 ml)"),
    ("15455542", "Carabins Gin Cerises Noires (4 x 355 ml)"),
    ("15469776", "Carabins Gin Limonade Fraise Kiwi (4 x 355 ml)"),
    ("15051880", "Kopparberg Fruits Tropicaux (4 x 500 ml)"),
    ("15162230", "Palm Bay Citrus Zero Sucre (4 x 355 ml)"),
    ("15455577", "Carabins Gin Miel Citron (4 x 355 ml)"),
    ("15298225", "Malibu Mangue (4 x 355 ml)"),
    ("15455614", "Carabins Gin Limonade Mangue (4 x 355 ml)"),
    ("15617396", "Twisted Tea Extra"),
    ("15498235", "Carabins Gin Raisin (4 x 355 ml)"),
    ("15458161", "Carabins Gin Limonade Canneberge (4 x 355 ml)"),
    ("15179138", "Carabins Gin Limonade Melon (4 x 355 ml)"),
    ("12028262", "Kopparberg Fraise et Lime sans alcool (4 x 500 ml)"),
    ("15299818", "Mike's Hard Lemonade Peche"),
    ("14960184", "Kopparberg Fraise et Lime 0% (4 x 500 ml)"),
    ("15455606", "Carabins Gin Limonade Tropical (4 x 355 ml)"),
    ("15311531", "Romeo's Gin Rose Limonade (4x355 ml)"),
    ("15298364", "Smirnoff Seltzer Mangue (4 x 355 ml)"),
    ("15148083", "Palm Bay Fraise Citron Lime (4 x 355 ml)"),
    ("15148260", "Palm Bay Framboise Peche Zero Sucre (4 x 355 ml)"),
    ("15148059", "Palm Bay Citron Lime Framboise (4 x 355 ml)"),
    ("14699473", "Chic Choc Rhum Epice et Citron (4 x 355 ml)"),
    ("15311549", "Romeo's Gin Limonade Fraise (4x355 ml)"),
    ("15124410", "Kopparberg Poire sans alcool (4 x 500 ml)"),
    ("15452270", "Carabins Gin Limonade Bleuet (4 x 355 ml)"),
    ("15302271", "Caisse Mixte Palm Bay (12 x 355 ml)"),
    ("15477590", "Carabins Gin Limonade Lime (4 x 355 ml)"),
    ("15049376", "Kopparberg Fruits des Bois sans alcool (4 x 500 ml)"),
    ("14701300", "Twisted Tea Hard Iced Tea Citron"),
    ("150920", "Mott's Clamato Caesar Classique Bouteille"),
    ("15298284", "Smirnoff Seltzer Framboise Sauvage (4 x 355 ml)"),
    ("15148067", "Palm Bay Framboise Citron Vert (4 x 355 ml)"),
    ("14961777", "Kopparberg Fraise sans alcool (4 x 500 ml)"),
    ("15149191", "Palm Bay Framboise Bleue Zero Sucre (4 x 355 ml)"),
    ("15455420", "Carabins Gin Fraise Banane (4 x 355 ml)"),
    ("15298348", "Smirnoff Seltzer Peche (4 x 355 ml)"),
    ("15149220", "Palm Bay Peche Mangue Zero Sucre (4 x 355 ml)"),
    ("15538735", "Carabins Gin Limonade Cerises (4 x 355 ml)"),
    ("15448166", "Carabins Gin Limonade Mure (4 x 355 ml)"),
    ("15149211", "Palm Bay Ananas Citron Vert Zero Sucre (4 x 355 ml)"),
    ("15299480", "Mike's Hard Lemonade Limonade Classique"),
    ("15021461", "Kopparberg Peche sans alcool (4 x 500 ml)"),
    ("15298022", "Malibu Citron (4 x 355 ml)"),
    ("11106468", "Mott's Clamato Caesar Classique (1.89 L)"),
    ("11440255", "Mott's Clamato Caesar Classique (946 ml)"),
    ("14649450", "Smirnoff Ice Limonade (4 x 330 ml)"),
    ("15538743", "Carabins Gin Limonade Framboise Bleue (4 x 355 ml)"),
    ("15350468", "Carabins Gin Limonade Ananas (4 x 355 ml)"),
    ("15298129", "Malibu Framboise (4 x 355 ml)"),
    ("14399331", "Palm Bay Citrus (4 x 355 ml)"),
    ("15125957", "Kopparberg Citron sans alcool (4 x 500 ml)"),
    ("15455489", "Carabins Gin Peche Abricot (4 x 355 ml)"),
    ("14960213", "Kopparberg Poire 0% (4 x 500 ml)"),
    ("15299455", "Twisted Tea Originale Bouteille 6 x 341 ml"),
    ("15078249", "Kopparberg Mangue sans alcool (4 x 500 ml)"),
    ("15200291", "Carabins Gin Limonade Pamplemousse (4 x 355 ml)"),
    ("14961793", "Kopparberg Fruits des Bois 0% (4 x 500 ml)"),
    ("15357281", "Carabins Gin Limonade Passion (4 x 355 ml)"),
    ("15298305", "Smirnoff Seltzer Framboise Sauvage (12 x 355ml)"),
    ("15149115", "Palm Bay Fraise Kiwi Zero Sucre (4 x 355 ml)"),
    ("15162184", "Palm Bay Peche Citron Zero Sucre (4 x 355 ml)"),
    ("15148534", "Palm Bay Framboise Bleue Citron Zero Sucre (4 x 355 ml)"),
    ("14699625", "Chic Choc Rhum Epice et Citron Zero Sucre (4 x 355 ml)"),
    ("15298428", "Smirnoff Seltzer Citron Vert (4 x 355 ml)"),
    ("15148219", "Palm Bay Fraise Ananas Zero Sucre (4 x 355 ml)"),
    ("14962041", "Kopparberg Fraise 0% (4 x 500 ml)"),
    ("14968119", "Kopparberg Peche et Framb 0% (4 x 500 ml)"),
    ("14960264", "Kopparberg Citron 0% (4 x 500 ml)"),
    ("14959802", "Kopparberg Mangue et Framb 0% (4 x 500 ml)"),
]

CACHE_FILE = "saq_id_cache.json"
ID_MAP = {
    "11927661": "203237",
    "14937996": "343304",
    "15298102": "388099",
    "14699633": "328078",
    "13991835": "278912",
    "15298807": "385077",
    "13990912": "278807",
    "542977": "295925",
    "15455340": "407382",
    "15298321": "384542",
    "14698016": "328105",
    "14536367": "316999",
    "15585814": "426886",
    "15455585": "406727",
    "15313499": "389772",
    "11927670": "203243",
    "15299463": "389385",
    "15148358": "364520",
    "14699035": "328132",
    "15372093": "394578",
    "14397993": "312776",
    "14969980": "346136",
    "15570831": "427144",
    "15557119": "420086",
    "15570794": "424262",
    "15565696": "424190",
    "15572714": "426527",
    "15372122": "394392",
    "15570823": "425337",
    "15580191": "426515",
    "13612137": "263129",
    "15162221": "367638",
    "15455382": "407994",
    "15570807": "424274",
    "13248646": "250883",
    "15585689": "426889",
    "15544570": "417396",
    "15455622": "407436",
    "11927741": "203261",
    "15570778": "425286",
    "10933998": "167876",
    "15103598": "359483",
    "15646541": "431282",
    "15298330": "385083",
    "15570760": "425147",
    "15148905": "364589",
    "15585726": "425685",
    "15455366": "407394",
    "15585865": "425697",
    "14699617": "328147",
    "14959562": "346000",
    "15472756": "408525",
    "15381926": "395346",
    "14701334": "329660",
    "15555121": "427726",
    "15298090": "385071",
    "15570858": "424280",
    "15148016": "364517",
    "14700235": "328153",
    "14398961": "312989",
    "15148032": "364571",
    "15299586": "385092",
    "14990949": "352139",
    "15585515": "427153",
    "13248638": "250880",
    "15585478": "426892",
    "15588935": "427156",
    "15372114": "394389",
    "15455307": "407997",
    "15404104": "399366",
    "14699561": "328066",
    "15298532": "385095",
    "15298479": "389355",
    "15570719": "424637",
    "15646550": "431285",
    "15455391": "407985",
    "15585777": "425361",
    "15570743": "424850",
    "15474217": "410679",
    "14961734": "346045",
    "15455411": "407406",
    "14959693": "345973",
    "15611584": "431357",
    "15296756": "391971",
    "15372106": "396978",
    "15455518": "406715",
    "15298372": "388102",
    "14398777": "312779",
    "15455551": "406721",
    "15599431": "426868",
    "14699596": "328144",
    "15298081": "388759",
    "15455462": "406703",
    "15162117": "366428",
    "15585873": "427150",
    "15148382": "364559",
    "15148649": "364568",
    "14700497": "328162",
    "15601589": "427162",
    "15586921": "426533",
    "15570735": "421601",
    "15585460": "425694",
    "15550419": "419045",
    "14398751": "313007",
    "15149094": "364514",
    "15585742": "427147",
    "15301261": "389391",
    "15148075": "364541",
    "15455374": "407400",
    "15455657": "407439",
    "15455358": "407391",
    "15585911": "426094",
    "15547607": "418516",
    "15148121": "364577",
    "15298014": "389373",
    "15585857": "425688",
    "15608342": "428920",
    "14016020": "280076",
    "15455294": "406327",
    "14968135": "346151",
    "15570840": "424268",
    "15299914": "385101",
    "15455631": "408015",
    "15313481": "389775",
    "15298006": "385068",
    "15455403": "408807",
    "15585582": "428005",
    "15586981": "427015",
    "15570903": "421745",
    "15299412": "390546",
    "15588353": "426880",
    "15455526": "406827",
    "15305579": "388111",
    "15570727": "421742",
    "15528730": "418066",
    "15586956": "427012",
    "15570751": "427141",
    "15455497": "406709",
    "15029412": "354778",
    "15637282": "429381",
    "14883553": "341817",
    "15587019": "426088",
    "15614857": "429922",
    "15585697": "428128",
    "15455446": "408012",
    "15585970": "426542",
    "15549389": "424430",
    "15471534": "408837",
    "15455534": "406718",
    "14559841": "319179",
    "14959773": "345931",
    "14300282": "304469",
    "15455438": "410106",
    "15299949": "385098",
    "15518321": "416331",
    "15020442": "355674",
    "14700121": "328087",
    "15077895": "358049",
    "15462793": "407481",
    "15149238": "364526",
    "15298356": "385089",
    "15586948": "428125",
    "15455649": "407433",
    "15149203": "367635",
    "15626292": "429573",
    "15148307": "364547",
    "15585620": "425271",
    "14700075": "328090",
    "15528625": "416898",
    "15525301": "414751",
    "14699588": "328141",
    "15455331": "407661",
    "15586999": "428008",
    "15498227": "412521",
    "15168156": "367632",
    "15545185": "424646",
    "15311485": "388282",
    "15455454": "409701",
    "15574277": "426898",
    "15126976": "366689",
    "15455315": "406700",
    "15574162": "424625",
    "14959491": "345946",
    "15498219": "412524",
    "15148139": "364550",
    "14889648": "341358",
    "15455471": "407664",
    "15055709": "357535",
    "15298313": "385086",
    "15148235": "364586",
    "15298188": "385107",
    "15455542": "407412",
    "15469776": "407496",
    "15051880": "356136",
    "15162230": "366644",
    "15455577": "406724",
    "15298225": "385104",
    "15455614": "407445",
    "15617396": "428717",
    "15498235": "412527",
    "15458161": "408546",
    "15179138": "367641",
    "12028262": "206174",
    "15299818": "389250",
    "14960184": "346006",
    "15455606": "408834",
    "15311531": "388279",
    "15298364": "385110",
    "15148083": "364532",
    "15148260": "364574",
    "15148059": "365696",
    "14699473": "328126",
    "15311549": "388276",
    "15124410": "362798",
    "15452270": "404788",
    "15302271": "389394",
    "15477590": "413850",
    "15049376": "355746",
    "14701300": "329186",
    "15298284": "389379",
    "15148067": "364553",
    "14961777": "345925",
    "15149191": "370605",
    "15455420": "413847",
    "15298348": "390543",
    "15149220": "364259",
    "15538735": "417420",
    "15448166": "406803",
    "15149211": "367629",
    "15299480": "387400",
    "15021461": "353096",
    "15298022": "388558",
    "11106468": "174284",
    "11440255": "185438",
    "14649450": "327350",
    "15538743": "417423",
    "15350468": "392658",
    "15298129": "387427",
    "14399331": "312998",
    "15125957": "363140",
    "15455489": "406706",
    "14960213": "345991",
    "15299455": "385731",
    "15078249": "358046",
    "14961793": "346027",
    "15357281": "395067",
    "15298305": "388945",
    "15149115": "364529",
    "15162184": "366425",
    "15148534": "364535",
    "14699625": "328060",
    "15298428": "385080",
    "15148219": "367494",
    "14962041": "346036",
    "14968119": "346145",
    "14960264": "346009",
    "14959802": "346021",
}
DELAI = 1.2

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "fr-CA,fr;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://www.saq.com/en/15544570",
    "Origin": "https://www.saq.com",
    "Connection": "keep-alive",
})

def load_cache():
    if Path(CACHE_FILE).exists():
        with open(CACHE_FILE, "r") as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)

def get_internal_id(code_saq, cache):
    if code_saq in ID_MAP:
        return ID_MAP[code_saq]
    if code_saq in cache:
        return cache[code_saq]
    return None

def get_distribution(internal_id):
    url = (
    f"https://www.saq.com/en/store/locator/ajaxlist/context/product/id/{internal_id}"
    f"?loaded=0&latitude=45.5088&longitude=-73.5540"
)
    try:
        r = SESSION.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()
        stores = data.get("list", [])
        total = data.get("total", len(stores))
        return {
            "total": total,
            "stores": [
                {
                    "nom": s.get("name", ""),
                    "ville": s.get("city", ""),
                    "qty": s.get("qty", 0),
                    "type": s.get("additional_attributes", {}).get("type", {}).get("label", ""),
                }
                for s in stores
            ],
            "erreur": "",
        }
    except Exception as e:
        return {"total": 0, "stores": [], "erreur": str(e)}

def build_excel(resultats, filepath):
    wb = openpyxl.Workbook()
    NAVY = "002060"
    GREY = "F2F2F2"
    WHITE = "FFFFFF"
    GREEN = "00703C"
    RED = "C00000"

    ws = wb.active
    ws.title = "Resume distribution"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Distribution SAQ — {datetime.today().strftime('%d %B %Y')}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", start_color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Code SAQ", "Produit", "# Succursales", "Statut", "Erreur"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(resultats, 3):
        fill = PatternFill("solid", start_color=GREY if i % 2 == 0 else WHITE)
        vals = [r["code"], r["nom"], r["total"], r["statut"], r["erreur"]]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center" if col_idx in (1, 3) else "left")
        nb = ws.cell(row=i, column=3)
        if r["total"] == 0:
            nb.font = Font(name="Arial", size=10, bold=True, color=RED)
        elif r["total"] >= 100:
            nb.font = Font(name="Arial", size=10, bold=True, color=GREEN)

    col_widths = [14, 45, 15, 12, 40]
    for idx in range(1, 6):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths[idx - 1]

    ws2 = wb.create_sheet("Detail succursales")
    headers2 = ["Code SAQ", "Produit", "Succursale", "Type", "Ville", "Qte"]
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center")

    row = 2
    for r in resultats:
        for s in r["stores"]:
            vals = [r["code"], r["nom"], s["nom"], s["type"], s["ville"], s["qty"]]
            for col_idx, val in enumerate(vals, 1):
                ws2.cell(row=row, column=col_idx, value=val).font = Font(name="Arial", size=9)
            row += 1

    col_widths2 = [14, 38, 30, 18, 20, 8]
    for idx in range(1, 7):
        ws2.column_dimensions[get_column_letter(idx)].width = col_widths2[idx - 1]

    wb.save(filepath)
    print(f"\nFichier sauvegarde : {filepath}")

def main():
    cache = load_cache()
    print(f"Cache existant : {len(cache)} ID(s) deja connus\n")

    resultats = []
    for i, (code, nom) in enumerate(PRODUITS, 1):
        print(f"[{i:02d}/{len(PRODUITS)}] {code} — {nom[:40]}")
        internal_id = get_internal_id(code, cache)
        if not internal_id:
            print(f"       X ID interne introuvable")
            resultats.append({"code": code, "nom": nom, "total": 0,
                               "statut": "erreur_id", "erreur": "ID interne non trouve", "stores": []})
            time.sleep(DELAI)
            continue
        print(f"       ID interne : {internal_id}", end=" -> ")
        dispo = get_distribution(internal_id)
        if dispo["erreur"]:
            print(f"X {dispo['erreur']}")
            statut = "erreur_api"
        else:
            print(f"{dispo['total']} succursale(s)")
            statut = "ok"
        resultats.append({
            "code": code, "nom": nom, "total": dispo["total"],
            "statut": statut, "erreur": dispo["erreur"], "stores": dispo["stores"],
        })
        time.sleep(DELAI)

    date_str = datetime.today().strftime("%Y%m%d")
    output = f"Distribution_SAQ_{date_str}.xlsx"
    build_excel(resultats, output)
    ok = sum(1 for r in resultats if r["statut"] == "ok")
    erreurs = sum(1 for r in resultats if r["statut"] != "ok")
    print(f"\nRecap : {ok} produits OK - {erreurs} erreur(s)")

if __name__ == "__main__":
    main()
